#!/usr/bin/env python
# coding: utf-8

"""Download your Social Bicycles (SoBi) route data and save it locally in various formats."""

__version__ = '0.4.4'
version = __version__

import csv
import datetime
import json
import logging
import openpyxl
import random
import requests
import time

from dicttoxml import dicttoxml
from openpyxl.writer.excel import save_virtual_workbook
from StringIO import StringIO

class Sobi(object):
    def __init__(self):
        self.username=''
        self.password=''
        self.path=''
        self.auth = None
        self.data = {
            'routes': [],
            'bikes': [],
            'hubs': [],
            'totals': {
                'total_distance_miles': 0,
                'total_distance_km': 0,
                'total_duration_seconds': 0,
                'total_duration_minutes': 0,
                'total_duration_hours': 0,
                'distinct_bikes': 0,
                'distinct_hubs': 0,
                'total_routes': 0,
            }
        }
        self.datetime_format = '%Y-%m-%dT%H:%M:%SZ'
        self.routes_url = 'https://app.socialbicycles.com/api/routes.json?page=%s'
        self.route_url = 'https://app.socialbicycles.com/api/routes/%s.json'
        self.hubs_url = 'https://app.socialbicycles.com/api/hubs.json'
        self.hub_url = 'https://app.socialbicycles.com/api/hubs/%s.json'
        self.bikes_url = 'https://app.socialbicycles.com/api/bikes.json'
        self.bike_url = 'https://app.socialbicycles.com/api/bikes/%s.json'
        self.friends_url = 'https://app.socialbicycles.com/api/friends.json'
        self.me_url = 'https://app.socialbicycles.com/api/users/me.json'
        self.polite = False

    def make_auth(self):
        if self.username == '':
            raise ValueError('The username cannot be an empty string')
        if self.password == '':
            raise ValueError('The password cannot be an empty string')
        self.auth = requests.auth.HTTPBasicAuth(self.username, self.password)

    def get_request(self, url):
        polite = self.polite
        auth=self.auth
        if not auth:
            self.make_auth()
        if polite == True: # wait before making the next query
            time.sleep(random.randrange(3)) 
        return requests.get(url, auth=auth)

    def get_data(self, page=1):
        if not self.auth:
            self.make_auth()
        this_routes_url = self.routes_url % (page)
        print('Getting %s...' % (this_routes_url))
        response = self.get_request(this_routes_url)
        if response.status_code != 200:
            self.calculate_totals()
            return
        routes_obj = response.json()
        items = routes_obj['items']
        if not items:
            self.calculate_totals()
            return
        for item in items:
            
            from_hub_address = self.lookup_hub_address(item['started_inside_hub_id'])
            to_hub_address = self.lookup_hub_address(item['finished_inside_hub_id'])
            distance_miles = self.to_float(item['distance'])
            distance_km = self.convert_miles_to_km(self.to_float(item['distance']))
            duration = self.duration_from_times(item['start_time'], item['finish_time'])
            duration_hh_mm_ss = self.duration_hh_mm_ss(duration)
            bike_name = self.lookup_bike_name(item['bike_id'])
            
            template = {
                'route_id': item['id'],
                'distance_miles': distance_miles,
                'distance_km': distance_km,
                'from_hub_id': item['started_inside_hub_id'],
                'from_hub_address': from_hub_address,
                'to_hub_id': item['finished_inside_hub_id'],
                'to_hub_address': to_hub_address,
                'start_time': item['start_time'],
                'finish_time': item['finish_time'],
                'duration': duration,
                'duration_hh_mm_ss': duration_hh_mm_ss,
                'first_location_address': item['first_location_address'],
                'bike_id': item['bike_id'],
                'bike_name': bike_name,
            }
            self.data['routes'].append(template)
        page = routes_obj['current_page'] + 1
        self.get_data(page=page)

    def lookup_bike_name(self, bike_id):
        if not bike_id:
            return ''
        bike_name = '' # initialize
        # check whether bike_id has already been looked up
        bike_match = [bike['bike_name'] for bike in self.data['bikes'] if bike['bike_id'] == bike_id]
        if not bike_match:
            this_bike_url = self.bike_url % (bike_id)
            response = self.get_request(this_bike_url)
            if response.status_code == 200:
                bike_obj = response.json()
                bike_name = bike_obj['name']
                # add bike to self.data['bikes']
                self.data['bikes'].append({ 'bike_id': bike_id, 'bike_name': bike_name })
        else: # bike_id has a match
            bike_name = bike_match[0]
        return bike_name

    def lookup_hub_address(self, hub_id):
        if not hub_id: # hub_id is None if route terminates outside a hub
            return ''
        hub_address = '' # initialize
        # check whether hub address has already been looked up
        hub_match = [hub['hub_address'] for hub in self.data['hubs'] if hub['hub_id'] == hub_id]
        if not hub_match:
            this_hub_url = self.hub_url % (hub_id)
            response = self.get_request(this_hub_url)
            if response.status_code == 200:
                hub_obj = response.json()
                hub_address = hub_obj['address']
                # add hub to self.data['hubs']
                self.data['hubs'].append({ 'hub_id': hub_id, 'hub_address': hub_address })
        else: # hub_id has a match
            hub_address = hub_match[0]
        return hub_address

    def duration_from_times(self, str_start_time, str_finish_time):
        start_time = datetime.datetime.strptime(str_start_time, self.datetime_format)
        finish_time = datetime.datetime.strptime(str_finish_time, self.datetime_format)
        delta = finish_time - start_time
        return int(delta.total_seconds())

    def duration_from_route(self, route):
        return route['duration']

    def total_duration_seconds(self, routes):
        return sum(self.duration_from_route(route) for route in routes)

    def total_duration_minutes(self, seconds):
        return round(seconds * 1.0 / 60, 2)

    def total_duration_hours(self, seconds):
        return round(seconds * 1.0 / 60 / 60, 2)

    def duration_hh_mm_ss(self, duration):
        return time.strftime('%H:%M:%S', time.gmtime(duration))

    def to_float(self, val):
        try:
            return float(val)
        except ValueError:
            return 0

    def distance_from_route(self, route):
        return self.to_float(route['distance_miles'])

    def total_distance(self, routes):
        return round(sum(self.distance_from_route(route) for route in routes), 2)

    def calculate_totals(self):
        self.data['totals']['total_distance_miles'] = self.total_distance(self.data['routes'])
        self.data['totals']['total_distance_km'] = self.convert_miles_to_km(self.data['totals']['total_distance_miles'])
        self.data['totals']['total_duration_seconds'] = self.total_duration_seconds(self.data['routes'])
        self.data['totals']['total_duration_minutes'] = self.total_duration_minutes(self.data['totals']['total_duration_seconds'])
        self.data['totals']['total_duration_hours'] = self.total_duration_hours(self.data['totals']['total_duration_seconds'])
        self.data['totals']['distinct_bikes'] = len(self.data['bikes'])
        self.data['totals']['distinct_hubs'] = len(self.data['hubs'])
        self.data['totals']['total_routes'] = len(self.data['routes'])

    def convert_miles_to_km(self, miles):
        return round(miles * 1.60934, 2)

    def get_sorted_keys(self, dictionary):
        return sorted(dictionary.keys())

    def save_data(self, ext='json'):
        ext = ext.lower()
        filename = '%s/sobidata_export.%s' % (self.path, ext)
        contents = self.export_data(self.data, ext)
        with open(filename, 'w') as myfile:
            myfile.write(contents)

    def export_data(self, data, ext='json'):
        if ext == 'json':
            return self.export_to_json(data)
        if ext == 'xml':
            return self.export_to_xml(data)
        if ext == 'csv':
            return self.export_to_csv(data)
        if ext == 'xlsx':
            return self.export_to_xlsx(data)
        raise ValueError('File extension %s is not supported for export' % (ext))
    
    def export_to_json(self, data):
        return json.dumps(data)

    def export_to_xml(self, data):
        return dicttoxml(data)

    def export_to_csv(self, data):
        # we're going to save the routes as a csv table
        routes = data['routes']
        # get the column headers
        fields = self.get_sorted_keys(routes[0])
        # open a file-like object for the csv module
        file_obj = StringIO()
        # create a csv writer
        csv_write = csv.writer(file_obj)
        # write the column headers
        csv_write.writerow(fields)
        # now add the rows
        for row in routes:
            csv_write.writerow([row[field] for field in fields])
        # extract the contents of the file-like object
        contents = file_obj.getvalue()
        # close the file-like object
        file_obj.close()
        # now return the contents
        return contents

    def export_to_xlsx(self, data):
        # we'll save each data group to its own tab
        routes = data['routes']
        hubs = data['hubs']
        bikes = data['bikes']
        # open a workbook
        wb = openpyxl.Workbook()
        # create the first sheet
        sheet1 = wb.active
        sheet1.title = 'Routes'
        # get the column headers
        fields = self.get_sorted_keys(routes[0])
        # write the column headers
        sheet1.append([field for field in fields])
        # write the rows
        for row in routes:
            sheet1.append([row[field] for field in fields])
        # create the second sheet
        sheet2 = wb.create_sheet(2, title='Hubs')
        # get the column headers
        fields = self.get_sorted_keys(hubs[0])
        # write the column headers
        sheet2.append([field for field in fields])
        # write the rows
        for row in hubs:
            sheet2.append([row[field] for field in fields])
        # create the third sheet
        sheet3 = wb.create_sheet(3, title='Bikes')
        # get the column headers
        fields = self.get_sorted_keys(bikes[0])
        # write the column headers
        sheet3.append([field for field in fields])
        # write the rows
        for row in bikes:
            sheet3.append([row[field] for field in fields])
        # create the fourth sheet
        sheet4 = wb.create_sheet(4, title='Totals')
        # get the totals
        fields = self.get_sorted_keys(data['totals'])
        # write the rows
        for field in fields:
            sheet4.append([field, data['totals'][field]])
        return save_virtual_workbook(wb)

    def import_data(self, ext='json'):
        filename = '%s.%s' % (self.path, ext)
        with open(filename, 'r') as myfile:
            content = myfile.read()
            if ext == 'json':
                self.data = json.loads(content)
            else:
                raise ValueError('File extension %s is not supported for import.' % (ext))

    def make_request(self, resource='me', item_id=None):
        """Execute a GET request against the SoBi API for a specific resource.
        Supports the following resources:
        * routes - the details for a route
        * hubs - the details for a hub
        * bikes - the details for a bike
        * friends - list of friends (no way to access individuals by id)
        * me - returns your own user details
        """
        if resource not in ('routes', 'hubs', 'bikes', 'friends', 'me'):
            raise ValueError('Resource %s is not supported.' % (resource))
        if resource == 'routes' and item_id:
            this_url = self.route_url % (item_id)
        elif resource == 'hubs' and item_id:
            this_url = self.hub_url % (item_id)
        elif resource == 'bikes' and item_id:
            this_url = self.bike_url % (item_id)
        elif resource == 'friends':
            this_url = self.friends_url % (1)
        elif resource == 'me':
            this_url = self.me_url
        else: # valid resource but no item_id
            raise ValueError('Resource %s is not currently supported without an item id.' % (resource))
        response = self.get_request(this_url)
        return response


